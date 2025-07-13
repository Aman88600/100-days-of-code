from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time
import os
from dotenv import load_dotenv
from openpyxl import load_workbook

# == getting env values ===
load_dotenv()

# === Paths ===
driver_path = os.getenv("DRIVER_PATH")
excel_path = os.getenv("EXCEL_PATH")
wb = load_workbook(excel_path)

# print(excel_path)


# === Load Excel file ===
wb = load_workbook(excel_path)
ws = wb.active

# === Destination to airport code mapping ===
airport_codes = {
    "Paris": "cdg",
    "Bali": "dps",
    "Jakarta": "cgk",
    "Kathmandu": "ktm",
    "Penang": "pen",
    "Colombo": "cmb",
    "Langkawi": "lgk",
    "Seoul": "icn",
    "Nairobi": "nbo",
    "Zurich": "zrh"
}

# === Set up Chrome driver ===
driver = webdriver.Chrome(service=Service(driver_path))

# === Flight price checker ===
def get_flight_price(destination):
    try:
        city = destination.split(",")[0].strip()
        code = airport_codes.get(city.lower().capitalize(), "")
        if not code:
            return "❌ No airport code"

        url = f"https://www.trip.com/flights/delhi-to-{city.lower()}/airfares-del-{code}/"
        driver.get(url)

        # Wait for the price to load
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "price-number"))
        )

        price_element = driver.find_element(By.CLASS_NAME, "price-number")
        price = int(price_element.text.replace("₹", "").replace(",", "").strip())
        return price

    except Exception as e:
        return f"❌ Error: {e}"

# === Main check ===
print("📊 Flight Price Comparison:\n")
for row in ws.iter_rows(min_row=2, values_only=True):
    destination, saved_price, _ = row
    current_price = get_flight_price(destination)

    if isinstance(current_price, int):
        if current_price < saved_price:
            trend = "🔻 Lowered"
        elif current_price > saved_price:
            trend = "🔼 Increased"
        else:
            trend = "➖ Unchanged"
        print(f"{destination}: ₹{saved_price} → ₹{current_price} | {trend}")
    else:
        print(f"{destination}: {current_price}")

driver.quit()
