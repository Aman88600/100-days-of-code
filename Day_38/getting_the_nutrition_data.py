import re
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "exercise_log.xlsx"

def parse_duration_exercises(sentence):
    activities = [
        "run", "ran", "walk", "walked", "cycle", "cycled",
        "jog", "jogged", "swim", "swam", "gym", "exercise", "workout", "yoga"
    ]
    sentence = sentence.lower()
    pattern = r"(?P<activity>" + "|".join(activities) + r")[^0-9]*?(?P<duration>\d+(?:\.\d+)?)\s*(?P<unit>minutes|minute|hours|hour|mins|hrs)"

    matches = re.finditer(pattern, sentence)
    results = []

    for match in matches:
        activity = match.group("activity")
        duration = float(match.group("duration"))
        unit = match.group("unit")

        # Normalize unit
        if unit in ["minute", "minutes", "mins"]:
            unit = "minutes"
        elif unit in ["hour", "hours", "hrs"]:
            unit = "hours"

        results.append({
            "activity": activity,
            "duration": duration,
            "unit": unit
        })

    return results

def save_to_excel(logs):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Exercise Log"
        ws.append(["Date", "Time", "Activity", "Duration", "Unit"])
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Exercise Log"]

    for entry in logs:
        now = datetime.now()
        ws.append([
            now.strftime("%Y-%m-%d"),
            now.strftime("%H:%M"),
            entry["activity"],
            entry["duration"],
            entry["unit"]
        ])
    wb.save(EXCEL_FILE)
    print("✅ Data saved to", EXCEL_FILE)

sentence = input("Enter Exercise : ")
# sentence = "I did gym for 1 hour and walked for 45 minutes"
logs = parse_duration_exercises(sentence)
print(logs)
save_to_excel(logs)
